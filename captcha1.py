import time
import openpyxl
import os
from cryptography.fernet import Fernet
from captcha.image import ImageCaptcha
import cv2
import secrets
import string


wb = openpyxl.load_workbook("PF.xlsx")
sh1 = wb['Sheet1']
def cls(): os.system('cls')


def register():
    cls()
    print("REGISTRATION")
    userName = input("Enter your Username :")
    r_max = sh1.max_row
    for i in range(1, r_max+1):
        if sh1.cell(i, 1).value == userName.strip():
            print("USERNAME ALREADY USED, \nPLEASE TRY ANOTHER USERNAME")
            time.sleep(2)
            register()
    ps = input("Enter your password :")
    eps = ps.encode('utf-8')
    key = Fernet.generate_key()
    f = Fernet(key)
    token = f.encrypt(eps)
    sh1.cell(r_max+1, 1, value=userName)
    sh1.cell(r_max+1, 2, value=token)
    sh1.cell(r_max+1, 3, value=key)
    wb.save("PF.xlsx")
    print("\nRegistered Successfully")
    time.sleep(2)
    exit(0)


def login():
    cls()
    print("LOGIN")
    userName = input("Enter your Username :")
    r_max2 = sh1.max_row
    count = 0
    for i in range(1, r_max2 + 1):
        if sh1.cell(i, 1).value == userName.strip():
            count = 1
            ps = input("Enter your password :")
            tempKey = sh1.cell(i, 3).value
            key = tempKey.encode('utf-8')
            f = Fernet(key)
            tempToken = sh1.cell(i, 2).value
            token = tempToken.encode()
            checkPs = f.decrypt(token).decode()
            if checkPs == ps:
                print("LOGIN SUCCESSFUL \nPROCEEDING TO CAPTCHA RECOGNITION")
                time.sleep(2)
                captcha()
            else:
                print("WRONG PASSWORD, PLEASE TRY AGAIN")
                time.sleep(3)
                login()
    if count != 1:
        print("USERNAME NOT FOUND")
        ch = input("\nPRESS 1 TO REGISTER OR 2 TO RE-LOGIN OR 3 to EXIT")
        if ch == '1':
            register()
        elif ch == '2':
            login()
        else:
            exit(0)


def captcha():
    cls()
    print("CAPTCHA RECOGNITION")
    print("Read the generated captcha and enter the text")
    captcha_text = str(''.join(secrets.choice(string.ascii_uppercase + string.digits)for i in range(5)))
    imj = ImageCaptcha(width=280, height=90)
    imj.generate(captcha_text)
    imj.write(captcha_text, "CAPTCHA.png")
    captcha_image = cv2.imread("CAPTCHA.png")
    cv2.imshow('CAPTCHA', captcha_image)
    cv2.waitKey(0)
    cv2.destroyAllWindows()
    ans = input("Enter the Text: ")
    if ans == captcha_text:
        print("LOGIN COMPLETE!!")
        time.sleep(2)
        exit(0)
    else:
        print("WRONG CAPTCHA ENTERED, \nPLEASE TRY AGAIN")
        time.sleep(2)
        captcha()

def main():
    cls()
    print("Encrypted Login System \n")
    choice = input("Press 1 to login \nPress 2 to Register \nPress anything else to exit")
    if choice == '1':
        login()
    elif choice == '2':
        register()
    else:
        print("Exiting the Program")
        time.sleep(2)
        exit(0)
    main()


if _name_ == "_main_":
    main()